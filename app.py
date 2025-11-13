from flask import Flask, render_template
from openpyxl import load_workbook
from pathlib import Path

APP_ROOT = Path(__file__).parent.resolve()
EXCEL_PATH = Path("Checklist.xlsx")

app = Flask(__name__, template_folder=str(APP_ROOT / "templates"), static_folder=str(APP_ROOT / "static"))

def html_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
         .replace('"', "&quot;")
         .replace("'", "&#39;")
    )

def to_html_with_style(cell):
    """Return cell value as HTML preserving bold/red font, including rich text."""
    v = cell.value
    if v is None:
        return ""
    
    # Check if cell has rich text (partial formatting)
    if hasattr(cell, 'font') and hasattr(cell, '_value') and hasattr(cell._value, '__iter__'):
        try:
            # Try to access rich text runs
            rich_text = cell._value
            if hasattr(rich_text, '__iter__') and not isinstance(rich_text, str):
                html_parts = []
                for run in rich_text:
                    if hasattr(run, 'text') and hasattr(run, 'font'):
                        text = html_escape(str(run.text))
                        # Check for bold
                        if getattr(run.font, 'bold', False):
                            text = f"<strong>{text}</strong>"
                        # Check for red color
                        color = getattr(run.font, 'color', None)
                        if color and getattr(color, 'rgb', None):
                            rgb = color.rgb.upper()
                            if rgb.endswith("FF0000") or rgb[-6:] == "FF0000":
                                text = f'<span class="red-text">{text}</span>'
                        html_parts.append(text)
                    elif isinstance(run, str):
                        html_parts.append(html_escape(run))
                if html_parts:
                    result = ''.join(html_parts)
                    # Handle line breaks
                    if "\n" in str(v) and len(str(v).split("\n")) <= 3 and len(str(v)) < 30:
                        result = result.replace("\n", " ")
                    else:
                        result = result.replace("\n", "<br>")
                    return result
        except:
            pass  # Fall back to regular processing
    
    # Regular processing (no rich text)
    raw = str(v)
    if "\n" in raw and len(raw.split("\n")) <= 3 and len(raw) < 30:
        text = html_escape(raw.replace("\n", " "))  # join with space
    else:
        text = html_escape(raw).replace("\n", "<br>")

    # Bold & red detection for entire cell
    is_bold = getattr(cell.font, "bold", False)
    color = getattr(cell.font, "color", None)
    is_red = False
    if color and getattr(color, "rgb", None):
        rgb = color.rgb.upper()
        if rgb.endswith("FF0000") or rgb[-6:] == "FF0000":
            is_red = True

    if is_bold:
        text = f"<strong>{text}</strong>"
    if is_red:
        text = f'<span class="red-text">{text}</span>'
    return text

def _merged_map(ws):
    top_left = {}
    covers = set()
    for mr in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        top_left[(min_row, min_col)] = (rowspan, colspan)
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                covers.add((rr, cc))
    return top_left, covers

def _row_has_warning(ws, r, merged_top_left):
    for (rr, cc), (rs, cs) in merged_top_left.items():
        if rr == r and cs >= 2:
            return True
    return False

def parse_workbook(path: Path):
    wb = load_workbook(filename=str(path), data_only=True)
    sheets = []

    for ws in wb.worksheets:
        merged_top_left, merged_covers = _merged_map(ws)
        rows = []

        for r in range(1, ws.max_row + 1):
            if _row_has_warning(ws, r, merged_top_left):
                warning_text = ""
                for (rr, cc), (rs, cs) in merged_top_left.items():
                    if rr == r and cs >= 2:
                        warning_text = to_html_with_style(ws.cell(row=rr, column=cc))
                        break
                rows.append({"type": "warning", "html": warning_text})
                continue

            cells = []
            c = 1
            # Skip the last column (Note column)
            max_col = ws.max_column - 1 if ws.max_column > 1 else ws.max_column
            
            while c <= max_col:
                if (r, c) in merged_covers:
                    span = merged_top_left.get((r, c))
                    if span:
                        rs, cs = span
                        # If merged cell extends into note column, reduce colspan
                        if c + cs - 1 > max_col:
                            cs = max_col - c + 1
                        html = to_html_with_style(ws.cell(row=r, column=c))
                        cells.append({"html": html, "rowspan": rs, "colspan": cs})
                        c += cs
                    else:
                        c += 1
                else:
                    html = to_html_with_style(ws.cell(row=r, column=c))
                    cells.append({"html": html, "rowspan": 1, "colspan": 1})
                    c += 1

            rows.append({"type": "data", "cells": cells})

        sheets.append({
            "name": ws.title,
            "rows": rows,
            "max_cols": max_col
        })

    return {"file_name": path.name, "sheets": sheets}

@app.route("/")
def index():
    model = parse_workbook(EXCEL_PATH)
    return render_template("index.html", model=model)

if __name__ == "__main__":
    app.run(debug=True)
