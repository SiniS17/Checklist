from pathlib import Path
from openpyxl import load_workbook
from formatters.cell_formatter import CellFormatter
from handlers.merge_handler import MergeHandler


class ExcelService:
    """Service for parsing Excel workbooks into HTML-ready data"""

    def __init__(self):
        self.formatter = CellFormatter()

    def parse_workbook(self, path: Path):
        """Parse entire workbook and return structured data"""
        wb = load_workbook(filename=str(path), data_only=True)
        sheets = []

        for ws in wb.worksheets:
            sheet_data = self._parse_worksheet(ws)
            sheets.append(sheet_data)

        return {
            "file_name": path.name,
            "sheets": sheets
        }

    def _parse_worksheet(self, worksheet):
        """Parse a single worksheet"""
        merge_handler = MergeHandler(worksheet)
        rows = []

        for r in range(1, worksheet.max_row + 1):
            # Check if this is a warning row
            if merge_handler.is_warning_row(r):
                warning_text = merge_handler.get_warning_text(r, self.formatter)
                rows.append({
                    "type": "warning",
                    "html": warning_text
                })
                continue

            # Parse regular data row
            row_data = self._parse_row(worksheet, r, merge_handler)
            rows.append(row_data)

        # Skip the last column (Note column)
        max_col = worksheet.max_column - 1 if worksheet.max_column > 1 else worksheet.max_column

        return {
            "name": worksheet.title,
            "rows": rows,
            "max_cols": max_col
        }

    def _parse_row(self, worksheet, row_num, merge_handler):
        """Parse a single row of cells"""
        cells = []
        col = 1
        max_col = worksheet.max_column - 1 if worksheet.max_column > 1 else worksheet.max_column

        while col <= max_col:
            if merge_handler.is_covered(row_num, col):
                span = merge_handler.get_span(row_num, col)

                if span:
                    rowspan, colspan = span

                    # If merged cell extends into note column, reduce colspan
                    if col + colspan - 1 > max_col:
                        colspan = max_col - col + 1

                    cell = worksheet.cell(row=row_num, column=col)
                    html = self.formatter.to_html_with_style(cell)

                    cells.append({
                        "html": html,
                        "rowspan": rowspan,
                        "colspan": colspan
                    })
                    col += colspan
                else:
                    col += 1
            else:
                cell = worksheet.cell(row=row_num, column=col)
                html = self.formatter.to_html_with_style(cell)

                cells.append({
                    "html": html,
                    "rowspan": 1,
                    "colspan": 1
                })
                col += 1

        return {
            "type": "data",
            "cells": cells
        }